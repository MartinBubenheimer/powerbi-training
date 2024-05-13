from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Load the existing workbook
wb = load_workbook("customer_sales_data_sheets.xlsx")

# Select the active worksheet
sheet_name = "April 2024"
ws = wb[sheet_name]

# Add column title to column I (9)
cell = ws.cell(row=1, column=9)
cell.value = "Sample Data"
cell.font = Font(bold=True)

# Generate 10 rows of data for column I (starting from row 2)
for i in range(1, 11):
    cell = ws[f'I{i + 1}']  # Offset by 1 to start from row 2
    cell.value = f"Data {i}"  # Insert sample data

# Save the workbook
wb.save("customer_sales_data_sheets.xlsx")